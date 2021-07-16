from openpyxl import load_workbook, drawing
import requests, datetime, re
from shutil import copyfile
from win32com import client
import pythoncom


def getCompanyInfo(id):
    try:
        res = requests.get('https://data.gcis.nat.gov.tw/od/data/api/5F64D864-61CB-4D0D-8AD9-492047CC1EA6?$format=json&$filter=Business_Accounting_NO%20eq%20'+id)
        return res.json()
    except:
        return '404'


def getQuotation(data):
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'}
    api_url = "https://data.gcis.nat.gov.tw/od/data/api/5F64D864-61CB-4D0D-8AD9-492047CC1EA6?$format=json&$filter=Business_Accounting_NO%20eq%2"
    datetime_dt = datetime.datetime.today()  # 獲得當地時間
    datetime_str = datetime_dt.strftime("%Y/%m/%d")  # 格式化日期

    wb = load_workbook('template.xlsx')
    sheet = wb['报价单']

    sheet['G3'] = datetime_str
    # deadline
    dateB = data.get('vday')
    sheet['G8'] = (datetime_dt + datetime.timedelta(days=int(dateB))).strftime("%Y/%m/%d")
    # name
    name = data.get('cname')
    sheet['B9'] = f'姓名：{name}'
    # 統編
    cId = data.get('taxid')
    if cId:
        res = getCompanyInfo(cId)[0]
        companyName, companyId, companyAddress = res['Company_Name'], res['Business_Accounting_NO'], res['Company_Location']
        sheet['B10'] = f'公司名稱：{companyName}'
        sheet['B11'] = f'統一編號：{companyId}'
        sheet['B12'] = f'公司地址：{companyAddress}'
    else:
        sheet['B11'] = f'統一編號：無'

    companyName, companyAddress = data.get('companyName'), data.get('companyAddress')
    if companyName:
        sheet['B10'] = f'公司名稱：{companyName}'
    if companyAddress:
        sheet['B12'] = f'公司地址：{companyAddress}'

    cPhone = data.get('cphone')
    sheet['B13'] = f'公司電話：'+cPhone

    ps = data.get('note')
    if ps:
        sheet['C16'] = ps
    else:
        sheet['C16'] = '無'

    pp = data.get('product').split('\r\n')
    products = [re.sub(r'[,|，]', '<sep>', p) for p in pp]
    products = [product.split('<sep>') for product in products]

    tax = data.get('tax')
    total = 0
    for i in range(24, 24+len(products)):
        sheet.merge_cells(f'C{i}:D{i}')
        sheet['B'+str(i)] = products[i-24][1]
        sheet['C'+str(i)] = products[i-24][0]
        sheet['E'+str(i)] = products[i-24][2]
        tmp = int(products[i-24][1]) * int(products[i-24][2])
        sheet['G' + str(i)] = tmp
        total += tmp
        if tax == 'n':
            sheet['F' + str(i)] = 'T'
        else:
            sheet['F' + str(i)] = ''
    sheet['G34'] = total
    sheet['G36'] = total * 0.05
    sheet['G38'] = total + total * 0.05

    seller = data.get('seller')
    sheet['B20'] = seller


    delivery = data.get('dday')
    if delivery:
        sheet['C20'] = delivery.replace('-', '/')
    else:
        sheet['C20'] = "-"

    delivery_way = data.get('delivery')
    sheet['D20'] = delivery_way

    cashflow = data.get('cash')
    sheet['F20'] = cashflow

    wb.save('output.xlsx')

    # to PDF
    fileName = res['Company_Name']+'報價單'+datetime_dt.strftime("%m%d")
    pythoncom.CoInitialize()
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False
    Workbook = app.Workbooks.Open('out.xlsx')
    try:
        Workbook.ActiveSheet.ExportAsFixedFormat(0, f'{fileName}.pdf')
    except Exception as e:
        print("Failed to convert in PDF format.Please confirm environment meets all the requirements and try again")
        print(str(e))
    finally:
        Workbook.Close(False)
        app.Quit()
    return fileName